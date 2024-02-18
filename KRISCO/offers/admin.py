from io import BytesIO, StringIO  # You already have these for handling in-memory files
import json  # Existing
import os
from django.urls import path, reverse  # Existing
from django.contrib import admin  # Existing
from django.shortcuts import get_object_or_404, render, redirect  # Existing
from django.http import HttpResponse, HttpResponseNotAllowed, HttpResponseRedirect, HttpResponseNotFound  # Existing, HttpResponse duplicated
from django.db import transaction  # Existing
import pandas as pd
from staff.models import StaffEmailConfiguration  # Existing
from .models import EmailLog, Weekly_Offer, BrandOffer  # Existing
import csv  # Existing
from django.db.models import Q  # Existing
import logging  # Existing
from django.contrib import messages  # Existing
from inventory.models import SlowMoversReport  # Existing
from .forms import DiscountForm, EditDiscountForm, EditSalonDiscountForm, SalonDiscountForm  # Existing
from django.template.defaultfilters import slugify  # Existing
from django.http import JsonResponse  # Existing
from daterange.filters import DateRangeFilter  # Existing
from openpyxl.utils import get_column_letter  # Existing
from openpyxl import Workbook  # Existing, ensure it's used if necessary
from django.utils import timezone  # Existing
from django.core.files.base import ContentFile  # Existing
from django.contrib.auth.decorators import login_required  # For user authentication in views
from django.contrib.auth.models import User  # If you're referencing the User model directly
from django.views.decorators.http import require_http_methods
from .forms import CustomerFilterForm
from customer.models import Customer
from django.views.decorators.csrf import csrf_exempt
from django.core.cache import cache
from django.views.decorators.http import require_POST
from django.core.mail import EmailMessage, get_connection

@admin.register(EmailLog)
class EmailLogAdmin(admin.ModelAdmin):
    list_display = ('recipient_email', 'subject', 'sent_attachment','status', 'sent_at', 'error_message')
    list_filter = ('status', 'sent_at')
    search_fields = ('recipient_email', 'subject', 'message')
    date_hierarchy = 'sent_at'
    readonly_fields = ('sent_at',)  # Makes the sent_at field read-only in the admin detail view

    def has_add_permission(self, request):
        # Optional: Disable the ability to add new logs via admin
        return False

    def has_delete_permission(self, request, obj=None):
        # Optional: Disable the ability to delete logs via admin
        return False


logger = logging.getLogger(__name__)

@admin.register(Weekly_Offer)
class Weekly_OfferAdmin(admin.ModelAdmin):
    list_display = ('sku', 'upc', 'description', 'brand', 'display_category', 'available_qty', 'msrp', 'discount', 'offer_price', 'required_quantity')
    search_fields = ['sku', 'description', 'brand', 'category']
    actions = ['download_csv_template']

    def display_category(self, obj):
        return obj.get_category_display()
    display_category.short_description = 'Category'
    

    def download_csv_template(self, request, queryset):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="Weekly_Offer_template.csv"'
        writer = csv.writer(response)
        headers = [field.name for field in Weekly_Offer._meta.get_fields() if field.name != 'id']
        writer.writerow(headers)
        return response

    download_csv_template.short_description = "Download CSV template for weekly offers"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-csv/', self.admin_site.admin_view(self.import_csv), name='import-csv'),
        ]
        return custom_urls + urls

    def import_csv(self, request):
            if request.method == 'POST':
                csv_file = request.FILES.get('csv_file')

                # Check if a file was uploaded
                if not csv_file:
                    messages.error(request, "No file was uploaded.")
                    return HttpResponseRedirect(request.path_info)

                # Check if the uploaded file is a CSV file
                if not csv_file.name.endswith('.csv'):
                    messages.error(request, "Invalid file format. Please upload a CSV file.")
                    return HttpResponseRedirect(request.path_info)

                try:
                    # Read and decode the file, then create a CSV reader object
                    file_data = csv_file.read().decode('utf-8')
                    csv_data = csv.reader(file_data.splitlines())

                    # Check if the header of the CSV matches the expected format
                    header = next(csv_data)
                    expected_header = ['sku', 'upc', 'description', 'brand', 'category', 'available_qty', 'msrp', 'discount', 'offer_price', 'required_quantity']
                    if header != expected_header:
                        messages.error(request, "Invalid CSV header.")
                        return HttpResponseRedirect(request.path_info)

                    # Process each row in the CSV
                    with transaction.atomic():
                        for row in csv_data:
                            obj, created = Weekly_Offer.objects.update_or_create(
                                sku=row[0],
                                defaults={
                                    'upc': int(row[1]),
                                    'description': row[2],
                                    'brand': row[3],
                                    'category': row[4],
                                    'available_qty': int(row[5]),
                                    'msrp': float(row[6]),
                                    'discount': float(row[7]),
                                    'offer_price': float(row[8]),
                                    'required_quantity': int(row[9])
                                }
                            )
                            if created:
                                logger.info(f"Created new Weekly_Offer: {obj.sku}")
                            else:
                                logger.info(f"Updated Weekly_Offer: {obj.sku}")

                    messages.success(request, "CSV file imported successfully.")
                    return HttpResponseRedirect('../')

                except Exception as e:
                    # Log any errors that occur during file processing
                    logger.error(f"Error processing CSV file: {e}")
                    messages.error(request, "An error occurred while processing the CSV file.")
                    return HttpResponseRedirect(request.path_info)

            # Render the CSV import form
            context = {
                'title': 'Import CSV for Weekly Offers',
            }
            return render(request, 'admin/import_csv.html', context)


logger = logging.getLogger(__name__)

def as_text(value):
    if value is None:
        return ""
    return str(value)
@admin.register(BrandOffer)
class BrandOfferAdmin(admin.ModelAdmin):
    list_display = ['date', 'time', 'offer_file','offer_type', 'created_person_first_name', 'created_person_last_name', 'created_person_email', 'customer_rank']
    list_filter = [('date',DateRangeFilter),'offer_type']  # Add more fields to filter as needed
    search_fields = ['offer_file','created_person_first_name', 'created_person_last_name', 'created_person_email', 'customer_rank']  # Add more fields to search as needed
    
    actions = ['generate_offers']

    def error(self,request):
        return render(request,'admin/offers/brandoffer/error.html')
    
    def success(self,request):
        return render(request, 'admin/offers/brandoffer/success.html')
    
    # Offer File Save Salon by Clicking the save salon button.
    
    def save_saloon(self,request):
        try:
            current_user = request.user

            # Retrieve the latest DataFrame from the session
            filtered_data_df_json = request.session.get('filtered_data_df')
            if not filtered_data_df_json:
                messages.error(request, 'No data found in session.')
                return HttpResponseRedirect('./')  # Adjust as needed

            # Load DataFrame from JSON
            df = pd.read_json(filtered_data_df_json, orient='split')

            # Ensure DataFrame is not empty and has required 'brand' column
            if df.empty or 'brand' not in df.columns:
                messages.error(request, 'The DataFrame is empty or missing the "brand" column.')
                return HttpResponseRedirect('./')  # Adjust as needed
            
            # Construct file name
            unique_brands = '_'.join(sorted(df['brand'].unique().tolist()))
            file_name = f"Krisco_{unique_brands}_{timezone.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
            # Drop specific columns and check for existence before dropping to avoid KeyError
            columns_to_drop = ['id', 'report_date', 'item_classification', 'qtyin_oneyear', 'qtyout_oneyear',
                            'balance_oneyear', 'available', 'cost','begining_balance', 'reference',
                            'percentage', 'sellercategory']
            df.drop(columns=[col for col in columns_to_drop if col in df.columns], inplace=True, errors='ignore')

            # Rename 'display_qty' to 'available' if 'display_qty' exists in DataFrame
            if 'display_qty' in df.columns:
                df.rename(columns={'display_qty': 'available'}, inplace=True)
            if 'salon' in df.columns:
                df.rename(columns={'salon' : 'cost'}, inplace=True)

            # Convert column names to uppercase
            df.columns = [col.upper() for col in df.columns]

            # Excel file creation
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')
                # Get workbook and worksheet for auto-adjusting column widths
                worksheet = writer.sheets['Sheet1']
                for column_cells in worksheet.columns:
                    length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells) + 2
                    worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length
                if 'UPC' in df.columns:
                    upc_col_idx = df.columns.get_loc('UPC') + 1  # +1 because DataFrame columns are 0-indexed but Excel columns are 1-indexed
                    for row in worksheet.iter_rows(min_col=upc_col_idx, max_col=upc_col_idx, min_row=2, max_row=worksheet.max_row):
                        for cell in row:
                            cell.number_format = '000000000000'

            # Prepare the file content for saving to the model's FileField
            output.seek(0)  # Rewind the buffer
            file_content = ContentFile(output.read(), name=file_name)

            with transaction.atomic():
                new_offer = BrandOffer(
                    date=timezone.now().date(),
                    time=timezone.now().time(),
                    offer_type='SALON',
                    created_by=current_user,
                    created_person_first_name=current_user.first_name,
                    created_person_last_name=current_user.last_name,
                    created_person_email=current_user.email,
                    customer_rank='DIAMOND'
                )
                new_offer.save()
                new_offer.offer_file.save(file_name, file_content, save=True)

            messages.success(request, 'Offer saved successfully.')
            return redirect('admin:index')  # Adjust as needed
        except Exception as e:
            logger.error(f"Error saving offer: {e}")
            messages.error(request, f"Error saving offer: {e}")
            return HttpResponseRedirect('admin:error')  # Adjust as needed
    
    # Offer File Save Salon by function call for email attachment.

    def save_offer_without_redirect_salon(self,request):
        try:
            current_user = request.user
            # Check if the DataFrame exists in the session
            filtered_data_df_json = request.session.get('filtered_data_df', '')
            if not filtered_data_df_json:
                messages.error(request, 'No data found in session.')
                return HttpResponse('No data found in session.')

            # Load the DataFrame from JSON
            df = pd.read_json(StringIO(filtered_data_df_json), orient='split')

            if df.empty or 'brand' not in df.columns:
                messages.error(request, 'The DataFrame is empty or missing the "brand" column.')
                return HttpResponse('The DataFrame is empty or missing the "brand" column.')

            # Processing steps
            unique_brands = '_'.join(sorted(df['brand'].unique().tolist()))
            file_name = f"Krisco_{unique_brands}_{timezone.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"

            # Drop specified columns if they exist to avoid KeyError
            columns_to_drop = ['id', 'report_date', 'item_classification', 'qtyin_oneyear', 'qtyout_oneyear',
                            'balance_oneyear', 'available', 'cost','begining_balance', 'reference',
                            'percentage', 'sellercategory']
            df.drop(columns=[col for col in columns_to_drop if col in df.columns], inplace=True, errors='ignore')

            # Rename 'display_qty' to 'available' if 'display_qty' exists in DataFrame
            if 'display_qty' in df.columns:
                df.rename(columns={'display_qty': 'available'}, inplace=True)
            if 'salon' in df.columns:
                df.rename(columns={'salon' : 'cost'},inplace=True)

            # Convert column names to uppercase
            df.columns = [col.upper() for col in df.columns]

            # Create Excel file
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')

                # Adjusting column widths
                worksheet = writer.sheets['Sheet1']
                for column_cells in worksheet.columns:
                    length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells) + 2
                    worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length

                # Formatting for UPC column
                if 'UPC' in df.columns:
                    upc_col_idx = df.columns.get_loc('UPC') + 1  # +1 for Excel's 1-indexing
                    for row in worksheet.iter_rows(min_col=upc_col_idx, max_col=upc_col_idx, min_row=2, max_row=worksheet.max_row + 1):
                        for cell in row:
                            cell.number_format = '000000000000'

            # Saving the file to the model's FileField
            output.seek(0)
            file_content = ContentFile(output.read(), name=file_name)

            # Database transaction
            with transaction.atomic():
                new_offer = BrandOffer(
                    date=timezone.now().date(),
                    time=timezone.now().time(),
                    offer_type='SALON',
                    created_by=current_user,
                    created_person_first_name=current_user.first_name,
                    created_person_last_name=current_user.last_name,
                    created_person_email=current_user.email,
                    customer_rank='DIAMOND'
                )
                new_offer.save()
                new_offer.offer_file.save(file_name, file_content, save=True)

            messages.success(request, 'Offer saved successfully.')
            return new_offer
        except Exception as e:
            messages.error(request, f"Error saving offer: {e}")
            return HttpResponse(f"Error saving offer: {e}")
        
    def commit_send_emails_salon(self, request):
        email_subject = "OFFER AUTOMATED EMAIL FROM KRISCO TEAM"
        email_body = "Hi please find the attached copy of the offer file."
        
        brand_offer_instance = self.save_offer_without_redirect_salon(request)
        if brand_offer_instance and hasattr(brand_offer_instance, 'offer_file'):
            file_path = brand_offer_instance.offer_file.path
            
            if 'customers_df' in request.session:
                customers_df_json = request.session['customers_df']
                df = pd.read_json(StringIO(customers_df_json), orient='split')
                
                all_emails_sent_successfully = True  # Initialize a flag to track overall success

                for index, row in df.iterrows():
                    staff_config = StaffEmailConfiguration.objects.filter(staff_id=row['staff_id']).first()
                    if not staff_config:
                        messages.error(request, f"No email configuration found for staff ID {row['staff_id']}.")
                        all_emails_sent_successfully = False
                        continue

                    cc_emails = row['customer_cc_email'].split(',') if row.get('customer_cc_email') else []
                    bcc_emails = row['customer_bcc_email'].split(',') if row.get('customer_bcc_email') else []

                    try:
                        connection = get_connection(
                            host=staff_config.host,
                            port=staff_config.port,
                            username=staff_config.username,
                            password=staff_config.password,
                            use_tls=staff_config.use_tls
                        )

                        email = EmailMessage(
                            email_subject,
                            email_body,
                            staff_config.username,
                            [row['email']],
                            cc=cc_emails,
                            bcc=bcc_emails,
                            connection=connection
                        )
                        email.attach_file(file_path)
                        email.send()

                        EmailLog.objects.create(
                            recipient_email=row['email'],
                            cc_emails=', '.join(cc_emails),
                            bcc_emails=', '.join(bcc_emails),
                            subject=email_subject,
                            message=email_body,
                            sent_attachment=os.path.basename(file_path),
                            status='Success',
                            error_message=''
                        )
                    except Exception as e:
                        EmailLog.objects.create(
                            recipient_email=row['email'],
                            cc_emails=', '.join(cc_emails),
                            bcc_emails=', '.join(bcc_emails),
                            subject=email_subject,
                            message=email_body,
                            sent_attachment=os.path.basename(file_path),
                            status='Failure',
                            error_message=str(e)
                        )
                        messages.error(request, f"Failed to send email to {row['email']}: {str(e)}")
                        all_emails_sent_successfully = False

                if all_emails_sent_successfully:
                    messages.success(request, 'All emails sent successfully.')
                    return redirect('admin:success')
                else:
                    messages.error(request, 'Some emails failed to send.')
                    return redirect(reverse('admin:error'))
        else:
            messages.error(request, 'Failed to get the brand offer instance or its file path.')
            return redirect(reverse('admin:error'))

        return render(request, 'admin/offers/brandoffer/send_offers_email_salon.html')


    def send_email_customers_salon(self, request, *args, **kwargs):
        if 'customers_df' in request.session:
            customers_df_json = request.session['customers_df']
            df = pd.read_json(StringIO(customers_df_json), orient='split')
            # Fetch staff details only if DataFrame is not empty
            staff_configs = StaffEmailConfiguration.objects.in_bulk(field_name='staff_id')
            # Add staff details to the DataFrame
            df['staff_first_name'] = df['staff_id'].apply(lambda x: staff_configs.get(x).first_name if x in staff_configs else '')
            df['staff_last_name'] = df['staff_id'].apply(lambda x: staff_configs.get(x).last_name if x in staff_configs else '')
            df['staff_email'] = df['staff_id'].apply(lambda x: staff_configs.get(x).username if x in staff_configs else '')
            df['customer_category'] = df['customer_category'].apply(lambda x: json.dumps(x) if isinstance(x, list) else x)
            # Convert DataFrame to list of dicts to pass to the template
            context = {'customers': df.to_dict(orient='records')}
            return render(request, 'admin/offers/brandoffer/send_offers_email_salon.html', context)
        else:
            messages.error(request, "No customer data found in session.")
            return redirect('admin:index')

    def email_customers_salon(self, request):
        # Check if the DataFrame is already in the session
        if 'customers_df' not in request.session:
            customers = Customer.objects.all().values()
            df = pd.DataFrame(list(customers))
            request.session['customers_df'] = df.to_json(orient='split')

        df_json = request.session.get('customers_df')

        if df_json == '[]':
            customers_list = []
        else:
            df = pd.read_json(StringIO(df_json), orient='split')

            # Fetch staff details only if DataFrame is not empty
            staff_configs = StaffEmailConfiguration.objects.in_bulk(field_name='staff_id')
            # Add staff details to the DataFrame
            df['staff_first_name'] = df['staff_id'].apply(lambda x: staff_configs.get(x).first_name if x in staff_configs else '')
            df['staff_last_name'] = df['staff_id'].apply(lambda x: staff_configs.get(x).last_name if x in staff_configs else '')
            df['staff_email'] = df['staff_id'].apply(lambda x: staff_configs.get(x).username if x in staff_configs else '')

            # Convert the customer_category column to a list of strings
            df['customer_category'] = df['customer_category'].apply(lambda x: json.dumps(x) if isinstance(x, list) else x)

            # Convert DataFrame to list of dicts to pass to the template
            customers_list = df.to_dict('records')

            # Get the unique customer categories
            customer_categories = df['customer_category'].unique().tolist()

        form = CustomerFilterForm()

        return render(request, 'admin/offers/brandoffer/email_customers_salon.html', {
            'customers': customers_list,
            'filter_form': form,
            'customer_categories': customer_categories
        })

    def reset_customers_df_salon(self,request):
        customers = Customer.objects.all().values()
        df = pd.DataFrame(list(customers))
        request.session['customers_df'] = df.to_json(orient='split')
        return redirect('admin:email_customers_salon')

    def remove_customer_salon(self,request, customer_id):
        if 'customers_df' in request.session:
            df_json = request.session.get('customers_df')
            df = pd.read_json(StringIO(df_json), orient='split')
            df = df[df['customer_id'] != customer_id]
            request.session['customers_df'] = df.to_json(orient='split')
            return redirect('admin:email_customers_salon')
        else:
            return redirect('admin:email_customers_salon')

    # Special Brand Offers Salon Views
    def edit_discount_salon_item(self,request, sku):
        response_data = {'success': False, 'message': '', 'data': {}}
        try:
            filtered_data_df_json = request.session.get('filtered_data_df')
            if not filtered_data_df_json:
                response_data['message'] = 'Session data not found.'
                return JsonResponse(response_data)

            df = pd.read_json(StringIO(filtered_data_df_json), orient='split')
            if sku not in df['sku'].values:
                response_data['message'] = 'Item not found.'
                return JsonResponse(response_data)

            if request.method == 'POST':
                form = EditSalonDiscountForm(request.POST)
                if form.is_valid():
                    idx = df.index[df['sku'] == sku].tolist()[0]
                    
                    # Update the DataFrame with form data
                    df.at[idx, 'description'] = form.cleaned_data['description']
                    df.at[idx, 'cost'] = float(form.cleaned_data['cost'])
                    
                    # Calculate salon as half of cost directly here instead of relying on form input for salon
                    salon_price = df.at[idx, 'cost'] / 2
                    df.at[idx, 'salon'] = salon_price
                    df.at[idx, 'display_qty'] = form.cleaned_data['display_qty']
                    df.at[idx, 'discount'] = float(form.cleaned_data['discount'])
                    df.at[idx, 'offer_price'] = round(salon_price * (1 - df.at[idx, 'discount'] / 100), 2)

                    request.session['filtered_data_df'] = df.to_json(orient='split')
                    response_data['success'] = True
                    response_data['message'] = 'Salon item updated successfully.'
                    response_data['data'] = df.loc[idx].to_dict()
                    return JsonResponse(response_data)
                else:
                    response_data['message'] = 'Form validation error.'
                    response_data['errors'] = form.errors.as_json()
                    return JsonResponse(response_data)
            else:
                response_data['message'] = 'Invalid request method.'
                return JsonResponse(response_data)
        except Exception as e:
            response_data['message'] = f'An error occurred: {str(e)}'
            return JsonResponse(response_data)

    def remove_discount_salon_item(self,request, sku):
        try:
            filtered_data_df_json = request.session.get('filtered_data_df')
            if not filtered_data_df_json:
                messages.error(request, 'Session data not found.')
                return redirect(request.META.get('HTTP_REFERER', 'redirect_if_referer_not_found'))

            df = pd.read_json(StringIO(filtered_data_df_json), orient='split')
            df = df[df['sku'] != sku]  # Remove the item
            request.session['filtered_data_df'] = df.to_json(orient='split')  # Update session

            messages.success(request, 'Item removed successfully.')
        except Exception as e:
            messages.error(request, f'An error occurred: {str(e)}')
        return redirect(request.META.get('HTTP_REFERER', 'redirect_if_referer_not_found'))

    def offer_discount_salon(self,request):
        context = {}
        filtered_data_df_json = request.session.get('filtered_data_df')

        # Initialize an empty DataFrame if session data is not found
        if filtered_data_df_json:
            df = pd.read_json(StringIO(filtered_data_df_json), orient='split')
            unique_brands = sorted(df['brand'].unique().tolist())
            df['display_qty'] = df['available']
            df['salon'] = df['cost'] / 2
        else:
            unique_brands = []
            df = pd.DataFrame(columns=['brand', 'cost', 'salon', 'discount', 'offer_price'])  # Define columns for consistency

        if request.method == 'POST':
            form = SalonDiscountForm(unique_brands, request.POST)
            if form.is_valid():
                cleaned_data = form.cleaned_data
                if not df.empty:
                    for brand in unique_brands:
                        field_name = f"additional_discount_{slugify(brand)}"
                        additional_discount_rate = cleaned_data.get(field_name, 0) / 100
                        df.loc[df['brand'] == brand, 'discount'] = round(additional_discount_rate * 100, 2)
                        df.loc[df['brand'] == brand, 'offer_price'] = round(df['salon'] * (1 - additional_discount_rate), 2)

                    request.session['filtered_data_df'] = df.to_json(orient='split')
                    messages.success(request, 'Salon offers updated successfully.')
                    # Reload the page with updated context to show changes immediately
                    return redirect(request.path)
                else:
                    messages.error(request, 'No data found to apply discounts to.')
            else:
                # Add form to context to display form errors
                context['form'] = form
        else:
            form = SalonDiscountForm(unique_brands)
        
        context['form'] = form
        if not df.empty:
            context['df'] = df.to_dict(orient='records')

        # Make sure to use the correct template path
        return render(request, 'admin/offers/brandoffer/offer_generation_salon.html', context)

    @staticmethod
    def offer_edit_salon(request):
        # This method needs to be able to handle the request object directly
        if 'filtered_data_df' in request.session:
            filtered_data_df_json = request.session['filtered_data_df']
            df = pd.read_json(StringIO(filtered_data_df_json), orient='split')

            # Prepare your context with the DataFrame
            context = {'df': df.to_dict(orient='records')}
            return render(request, 'admin/offers/brandoffer/offer_edit_salon.html', context)
        else:
            # If no DataFrame is found in the session, redirect or show an error
            messages.error(request, "No data found in session.")
            return redirect('admin:index')  # Adjust the redirect as needed

#=======================================SALON ALL CODES ABOVE==========================================
       
    def commit_send_emails(self, request):
        email_subject = "OFFER AUTOMATED EMAIL FROM KRISCO TEAM"
        email_body = "Hi please find the attached copy of the offer file."

        brand_offer_instance = self.save_offer_without_redirect(request)
        if brand_offer_instance and hasattr(brand_offer_instance, 'offer_file'):
            file_path = brand_offer_instance.offer_file.path
            
            if 'customers_df' in request.session:
                customers_df_json = request.session['customers_df']
                df = pd.read_json(StringIO(customers_df_json), orient='split')
                
                all_emails_sent_successfully = True

                for index, row in df.iterrows():
                    staff_config = StaffEmailConfiguration.objects.filter(staff_id=row['staff_id']).first()
                    if not staff_config:
                        messages.error(request, f"No email configuration found for staff ID {row['staff_id']}.")
                        all_emails_sent_successfully = False
                        continue

                    # Prepare lists for CC and BCC emails
                    cc_emails = [email.strip() for email in row['customer_cc_email'].split(',')] if row.get('customer_cc_email') else []
                    bcc_emails = [email.strip() for email in row['customer_bcc_email'].split(',')] if row.get('customer_bcc_email') else []

                    try:
                        connection = get_connection(
                            host=staff_config.host,
                            port=staff_config.port,
                            username=staff_config.username,
                            password=staff_config.password,
                            use_tls=staff_config.use_tls
                        )

                        email = EmailMessage(
                            email_subject,
                            email_body,
                            staff_config.username,
                            [row['email']],
                            cc=cc_emails,
                            bcc=bcc_emails,
                            connection=connection
                        )
                        email.attach_file(file_path)
                        email.send()

                        EmailLog.objects.create(
                            recipient_email=row['email'],
                            cc_emails=', '.join(cc_emails),
                            bcc_emails=', '.join(bcc_emails),
                            subject=email_subject,
                            message=email_body,
                            sent_attachment=os.path.basename(file_path),
                            status='Success',
                            error_message=''
                        )
                    except Exception as e:
                        EmailLog.objects.create(
                            recipient_email=row['email'],
                            cc_emails=', '.join(cc_emails),
                            bcc_emails=', '.join(bcc_emails),
                            subject=email_subject,
                            message=email_body,
                            sent_attachment=os.path.basename(file_path),
                            status='Failure',
                            error_message=str(e)
                        )
                        messages.error(request, f"Failed to send email to {row['email']}: {str(e)}")
                        all_emails_sent_successfully = False

                if all_emails_sent_successfully:
                    messages.success(request, 'All emails sent successfully.')
                    return redirect('admin:success')
                else:
                    messages.error(request, 'Some emails failed to send.')
                    return redirect(reverse('admin:error'))
        else:
            messages.error(request, 'Failed to get the brand offer instance or its file path.')
            return redirect(reverse('admin:error'))

        return render(request, 'admin/offers/brandoffer/send_offers_email.html')
  
    def save_offer_without_redirect(self,request):
        try:
            current_user = request.user
            # Check if the DataFrame exists in the session
            filtered_data_df_json = request.session.get('filtered_data_df', '')
            if not filtered_data_df_json:
                messages.error(request, 'No data found in session.')
                return HttpResponse('No data found in session.')

            # Load the DataFrame from JSON
            df = pd.read_json(StringIO(filtered_data_df_json), orient='split')

            if df.empty or 'brand' not in df.columns:
                messages.error(request, 'The DataFrame is empty or missing the "brand" column.')
                return HttpResponse('The DataFrame is empty or missing the "brand" column.')

            # Processing steps
            unique_brands = '_'.join(sorted(df['brand'].unique().tolist()))
            file_name = f"Krisco_{unique_brands}_{timezone.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"

            # Drop specified columns if they exist to avoid KeyError
            columns_to_drop = ['id', 'report_date', 'item_classification', 'qtyin_oneyear', 'qtyout_oneyear',
                            'balance_oneyear', 'available', 'begining_balance', 'reference',
                            'percentage', 'sellercategory']
            df.drop(columns=[col for col in columns_to_drop if col in df.columns], inplace=True, errors='ignore')

            # Rename column if exists
            if 'display_qty' in df.columns:
                df.rename(columns={'display_qty': 'available'}, inplace=True)

            # Convert column names to uppercase
            df.columns = [col.upper() for col in df.columns]

            # Create Excel file
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')

                # Adjusting column widths
                worksheet = writer.sheets['Sheet1']
                for column_cells in worksheet.columns:
                    length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells) + 2
                    worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length

                # Formatting for UPC column
                if 'UPC' in df.columns:
                    upc_col_idx = df.columns.get_loc('UPC') + 1  # +1 for Excel's 1-indexing
                    for row in worksheet.iter_rows(min_col=upc_col_idx, max_col=upc_col_idx, min_row=2, max_row=worksheet.max_row + 1):
                        for cell in row:
                            cell.number_format = '000000000000'

            # Saving the file to the model's FileField
            output.seek(0)
            file_content = ContentFile(output.read(), name=file_name)

            # Database transaction
            with transaction.atomic():
                new_offer = BrandOffer(
                    date=timezone.now().date(),
                    time=timezone.now().time(),
                    offer_type='REGULAR',
                    created_by=current_user,
                    created_person_first_name=current_user.first_name,
                    created_person_last_name=current_user.last_name,                    
                    created_person_email=current_user.email,
                    customer_rank='DIAMOND',
                )
                new_offer.save()
                new_offer.offer_file.save(file_name, file_content, save=True)

            messages.success(request, 'Offer saved successfully.')
            return new_offer
        except Exception as e:
            messages.error(request, f"Error saving offer: {e}")
            return HttpResponse(f"Error saving offer: {e}")

    def send_email_customers(self, request, *args, **kwargs):
        if 'customers_df' in request.session:
            customers_df_json = request.session['customers_df']
            df = pd.read_json(StringIO(customers_df_json), orient='split')
            # Fetch staff details only if DataFrame is not empty
            staff_configs = StaffEmailConfiguration.objects.in_bulk(field_name='staff_id')
            # Add staff details to the DataFrame
            df['staff_first_name'] = df['staff_id'].apply(lambda x: staff_configs.get(x).first_name if x in staff_configs else '')
            df['staff_last_name'] = df['staff_id'].apply(lambda x: staff_configs.get(x).last_name if x in staff_configs else '')
            df['staff_email'] = df['staff_id'].apply(lambda x: staff_configs.get(x).username if x in staff_configs else '')

            df['customer_category'] = df['customer_category'].apply(lambda x: json.dumps(x) if isinstance(x, list) else x)

            # Convert DataFrame to list of dicts to pass to the template
            context = {'customers': df.to_dict(orient='records')}
            return render(request, 'admin/offers/brandoffer/send_offers_email.html', context)
        else:
            messages.error(request, "No customer data found in session.")
            return redirect('admin:index')

    def email_customers(self, request):
        if 'customers_df' not in request.session:
            customers = Customer.objects.all().values()
            df = pd.DataFrame(list(customers))
            request.session['customers_df'] = df.to_json(orient='split')
        df_json = request.session.get('customers_df')

        if df_json == '[]':
            customers_list = []
        else:
            df = pd.read_json(StringIO(df_json), orient='split')
            staff_configs = StaffEmailConfiguration.objects.in_bulk(field_name='staff_id')
            df['staff_first_name'] = df['staff_id'].apply(lambda x: staff_configs.get(x).first_name if x in staff_configs else '')
            df['staff_last_name'] = df['staff_id'].apply(lambda x: staff_configs.get(x).last_name if x in staff_configs else '')
            df['staff_email'] = df['staff_id'].apply(lambda x: staff_configs.get(x).username if x in staff_configs else '')
            df['customer_category'] = df['customer_category'].apply(lambda x: json.dumps(x) if isinstance(x, list) else x)
            customers_list = df.to_dict('records')
            customer_categories = df['customer_category'].unique().tolist()

        form = CustomerFilterForm()

        return render(request, 'admin/offers/brandoffer/email_customers.html', {
            'customers': customers_list,
            'filter_form': form,
            'customer_categories': customer_categories
        })

    def reset_customers_df(self,request):
        customers = Customer.objects.all().values()
        df = pd.DataFrame(list(customers))
        request.session['customers_df'] = df.to_json(orient='split')
        return redirect('admin:email_customers')

    def remove_customer(self,request, customer_id):
        if 'customers_df' in request.session:
            df_json = request.session.get('customers_df')
            df = pd.read_json(StringIO(df_json), orient='split')
            df = df[df['customer_id'] != customer_id]
            request.session['customers_df'] = df.to_json(orient='split')
            return redirect('admin:email_customers')
#===============================Emailing Customer Views Above=======================================
    # Special Brand Offers Views
    def save_offer(self,request):
        try:
            current_user = request.user
            # Retrieve the latest DataFrame from the session
            filtered_data_df_json = request.session.get('filtered_data_df')
            if not filtered_data_df_json:
                messages.error(request, 'No data found in session.')
                return HttpResponseRedirect('./')  # Adjust as needed

            # Load DataFrame from JSON
            df = pd.read_json(filtered_data_df_json, orient='split')

            # Ensure DataFrame is not empty and has required 'brand' column
            if df.empty or 'brand' not in df.columns:
                messages.error(request, 'The DataFrame is empty or missing the "brand" column.')
                return HttpResponseRedirect('./')  # Adjust as needed
            
            # Construct file name
            unique_brands = '_'.join(sorted(df['brand'].unique().tolist()))
            file_name = f"Krisco_{unique_brands}_{timezone.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
            # Drop specific columns and check for existence before dropping to avoid KeyError
            columns_to_drop = ['id', 'report_date', 'item_classification', 'qtyin_oneyear', 'qtyout_oneyear',
                            'balance_oneyear', 'available', 'begining_balance', 'reference',
                            'percentage', 'sellercategory']
            df.drop(columns=[col for col in columns_to_drop if col in df.columns], inplace=True, errors='ignore')

            # Rename 'display_qty' to 'available' if 'display_qty' exists in DataFrame
            if 'display_qty' in df.columns:
                df.rename(columns={'display_qty': 'available'}, inplace=True)

            # Convert column names to uppercase
            df.columns = [col.upper() for col in df.columns]

            # Excel file creation
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')
                # Get workbook and worksheet for auto-adjusting column widths
                worksheet = writer.sheets['Sheet1']
                for column_cells in worksheet.columns:
                    length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells) + 2
                    worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length
                if 'UPC' in df.columns:
                    upc_col_idx = df.columns.get_loc('UPC') + 1  # +1 because DataFrame columns are 0-indexed but Excel columns are 1-indexed
                    for row in worksheet.iter_rows(min_col=upc_col_idx, max_col=upc_col_idx, min_row=2, max_row=worksheet.max_row):
                        for cell in row:
                            cell.number_format = '000000000000'

            # Prepare the file content for saving to the model's FileField
            output.seek(0)  # Rewind the buffer
            file_content = ContentFile(output.read(), name=file_name)

            with transaction.atomic():
                new_offer = BrandOffer(
                    date=timezone.now().date(),
                    time=timezone.now().time(),
                    offer_type='REGULAR',
                    created_by=current_user,
                    created_person_first_name=current_user.first_name,
                    created_person_last_name=current_user.last_name,
                    created_person_email=current_user.email,
                    customer_rank='DIAMOND'
                )
                new_offer.save()
                new_offer.offer_file.save(file_name, file_content, save=True)

            messages.success(request, 'Offer saved successfully.')
            return HttpResponseRedirect('/admin/offers/brandoffer/offer_save.html')  # Adjust as needed
        except Exception as e:
            logger.error(f"Error saving offer: {e}")
            messages.error(request, f"Error saving offer: {e}")
            return HttpResponseRedirect('/admin/offers/brandoffer/error.html')  # Adjust as needed

    def edit_discount_item(self, request, sku):
        response_data = {'success': False, 'message': '', 'data': {}}
        try:
            filtered_data_df_json = request.session.get('filtered_data_df')
            if not filtered_data_df_json:
                response_data['message'] = 'Session data not found.'
                return JsonResponse(response_data)

            df = pd.read_json(StringIO(filtered_data_df_json), orient='split')
            if sku not in df['sku'].values:
                response_data['message'] = 'Item not found.'
                return JsonResponse(response_data)

            if request.method == 'POST':
                form = EditDiscountForm(request.POST)
                if form.is_valid():
                    idx = df.index[df['sku'] == sku].tolist()[0]
                    df.at[idx, 'description'] = form.cleaned_data['description']
                    df.at[idx, 'display_qty'] = form.cleaned_data['display_qty']
                    df.at[idx, 'cost'] = float(form.cleaned_data['cost'])
                    df.at[idx, 'discount'] = float(form.cleaned_data['discount'])
                    df.at[idx, 'offer_price'] = round(float(form.cleaned_data['cost']) * (1 - float(form.cleaned_data['discount']) / 100), 2)

                    request.session['filtered_data_df'] = df.to_json(orient='split')
                    response_data['success'] = True
                    response_data['message'] = 'Item updated successfully.'
                    return JsonResponse(response_data)
                else:
                    response_data['message'] = 'Form validation error.'
                    return JsonResponse(response_data)
            else:
                response_data['message'] = 'Invalid request method.'
                return JsonResponse(response_data)
        except Exception as e:
            response_data['message'] = f'An error occurred: {str(e)}'
            return JsonResponse(response_data)
    
    def remove_discount_item(self, request, sku):
        try:
            filtered_data_df_json = request.session.get('filtered_data_df')
            if not filtered_data_df_json:
                messages.error(request, 'Session data not found.')
                return redirect(request.META.get('HTTP_REFERER', 'redirect_if_referer_not_found'))

            df = pd.read_json(StringIO(filtered_data_df_json), orient='split')
            
            # Logic to remove item based on sku
            df = df[df['sku'] != sku]
            request.session['filtered_data_df'] = df.to_json(orient='split')

            messages.success(request, 'Item removed successfully.')
        except Exception as e:
            messages.error(request, f'An error occurred: {str(e)}')
        return redirect(request.META.get('HTTP_REFERER', 'redirect_if_referer_not_found'))

    def offer_discount(self, request):
        context = {}
        filtered_data_df_json = request.session.get('filtered_data_df')
        if filtered_data_df_json:
            df = pd.read_json(StringIO(filtered_data_df_json), orient='split')
            unique_brands = sorted(df['brand'].unique().tolist())
            df['display_qty'] = df['available']
        else:
            unique_brands = []  # Fallback if there's no DataFrame data

        if request.method == 'POST':
            form = DiscountForm(unique_brands, request.POST)  # Initialize form with unique brands
            if form.is_valid():
                cleaned_data = form.cleaned_data
                if filtered_data_df_json:
                    # Loop through each dynamically created field and apply the discount
                    for brand in unique_brands:
                        field_name = f"discount_{slugify(brand)}"
                        discount_rate = cleaned_data.get(field_name, 0) / 100
                        # Round the discount rate to 2 decimals and apply it
                        df.loc[df['brand'] == brand, 'discount'] = round(discount_rate * 100, 2)  # Store as a percentage, rounded
                        df.loc[df['brand'] == brand, 'offer_price'] = (df['cost'] * (1 - discount_rate)).round(2)
                    # Save the updated DataFrame back to the session
                    request.session['filtered_data_df'] = df.to_json(orient='split')
                    context['df'] = df.to_dict(orient='records')  # Prepare data for template rendering
                else:
                    context['error'] = 'No data found to apply discounts to.'
            else:
                context['form'] = form
        else:
            form = DiscountForm(unique_brands)  # Initialize form with 'unique_brands' for GET requests
            context['form'] = form

            # Prepare existing DataFrame data for template rendering, if available
            if filtered_data_df_json:
                context['df'] = df.to_dict(orient='records')

        return render(request, 'admin/offers/brandoffer/offer_discount.html', context)

    @staticmethod
    def offer_edit(request):
        # This method needs to be able to handle the request object directly
        if 'filtered_data_df' in request.session:
            filtered_data_df_json = request.session['filtered_data_df']
            df = pd.read_json(StringIO(filtered_data_df_json), orient='split')

            # Prepare your context with the DataFrame
            context = {'df': df.to_dict(orient='records')}
            return render(request, 'admin/offers/brandoffer/offer_edit.html', context)
        else:
            # If no DataFrame is found in the session, redirect or show an error
            messages.error(request, "No data found in session.")
            return redirect('admin:index')  # Adjust the redirect as needed

#======================================Regular Offer Views Above====================================

    # Brand Offers Generation Views
    def generate_offers(self, request, queryset):
        # Retrieve selected filters from session storage
        selected_filters_json = request.session.get('selectedFilters')
        selected_filters = json.loads(selected_filters_json) if selected_filters_json else {}

        value = SlowMoversReport.objects.all()
        unique_brands = sorted(set(record.brand for record in value))
        unique_sellercategories = sorted(set(record.sellercategory for record in value))

        # Store selected brands in session
        selected_filters['selected_brands'] = request.POST.getlist('brand', [])
        request.session['selectedFilters'] = json.dumps(selected_filters)

        return render(request, 'admin/offers/brandoffer/offer_generation.html', {'unique_brands': unique_brands, 'unique_sellercategories': unique_sellercategories, 'selected_filters': selected_filters})

    def filter_offers(self, request, queryset=None):
            if request.method == 'POST':
                sellercategory = request.POST.getlist('sellercategory', [])
                brands = request.POST.getlist('brand', [])

                # Create a list to hold the brand filters
                brand_filters = []

                # Loop through selected brands and create a Q object for each
                for selected_brand in brands:
                    brand_filter = Q(brand__startswith=selected_brand)
                    brand_filters.append(brand_filter)

                # Combine the brand filters using the | operator
                combined_brand_filters = Q()
                for brand_filter in brand_filters:
                    combined_brand_filters |= brand_filter

                # Filter records based on user input and exclude records where 'available' is 0
                filtered_data = SlowMoversReport.objects.filter(
                    Q(sellercategory__in=sellercategory),
                    combined_brand_filters,
                    available__gt=0  # Greater than 0
                )

                # Create a DataFrame 'df' from the 'filtered_data'
                df = pd.DataFrame(list(filtered_data.values()))
                # Store the DataFrame in the session
                request.session['filtered_data_df'] = df.to_json(orient='split')

                # Render the filtered data in a new template
                return render(request, 'admin/offers/brandoffer/filtered_data.html', {'filtered_data': filtered_data})

            return render(request, 'admin/offers/brandoffer/offer_generation.html', {'value': queryset}) 
#========================Brand Offer Intial Generation and Filtering=====================

    generate_offers.short_description = 'Generate Brand Offers'

    # Paths for Views Handles Brand Offers
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            #error page
            path('error/',self.admin_site.admin_view(self.error),name='error'),
            #paths
            path('generate_offers/', self.admin_site.admin_view(self.generate_offers), name='generate_offers'),
            path('filter_offers/', self.admin_site.admin_view(self.filter_offers), name='filter_offers'),
            path('offer_discount/', self.admin_site.admin_view(self.offer_discount), name='offer_discount'),
            path('offer_edit/', self.admin_site.admin_view(self.offer_edit), name='admin_offer_edit'),
            path('offer_edit/edit/<str:sku>/', self.admin_site.admin_view(self.edit_discount_item), name='edit_discount_item'),
            path('offer_edit/remove/<str:sku>/', self.admin_site.admin_view(self.remove_discount_item), name='remove_discount_item'),
            # Ensure the URL for offer_edit_salon is for viewing the salon offers, not editing or removing directly
            path('offer_edit_salon/', self.admin_site.admin_view(self.offer_edit_salon), name='admin_offer_salon'),
            path('offer_discount_salon/', self.admin_site.admin_view(self.offer_discount_salon), name='offer_discount_salon'),
            # Correct the paths for editing and removing salon items to match the salon offer editing and removing operations
            path('offer_edit_salon/edit/<str:sku>/', self.admin_site.admin_view(self.edit_discount_salon_item), name='edit_discount_salon_item'),
            path('offer_edit_salon/remove/<str:sku>/', self.admin_site.admin_view(self.remove_discount_salon_item), name='remove_discount_salon_item'),
            # Save Brand Offer Salon
            path('save_salon/',self.admin_site.admin_view(self.save_saloon),name='save_salon'),
            path('save_offer/',self.admin_site.admin_view(self.save_offer), name='save_offer'),
            #Load customer Data
            path('email_customers/reset/', self.admin_site.admin_view(self.reset_customers_df), name='reset_customers_df'),
            path('email_customers/reset_salon/', self.admin_site.admin_view(self.reset_customers_df_salon), name='reset_customers_df_salon'),
            # Remove Email
            path('remove_customer/<str:customer_id>/', self.admin_site.admin_view(self.remove_customer), name='remove_customer'),
            path('remove_customer_salon/<str:customer_id>/', self.admin_site.admin_view(self.remove_customer_salon), name='remove_customer_salon'),
            # Email Brand and Salon Offer to Customer 
            path('email_customers/',self.admin_site.admin_view(self.email_customers),name='email_customers'),
            path('email_customers_salon/',self.admin_site.admin_view(self.email_customers_salon),name='email_customers_salon'),
            # Send Email 
            path('send_email_customers/',self.admin_site.admin_view(self.send_email_customers),name='send_email_customers'),
            path('send_email_customers_salon/',self.admin_site.admin_view(self.send_email_customers_salon),name='send_email_customers_salon'),
            # Save Without Redirection 
            path('save_offer_without_redirect/',self.admin_site.admin_view(self.save_offer_without_redirect),name='save_offer_without_redirect'),
            path('save_offer_without_redirect_salon/',self.admin_site.admin_view(self.save_offer_without_redirect_salon),name='save_offer_without_redirect_salon'),
            # Commit Send Emails
            path('commit_send_emails/',self.admin_site.admin_view(self.commit_send_emails),name='commit_send_emails'),
            path('commit_send_emails_salon/',self.admin_site.admin_view(self.commit_send_emails_salon),name='commit_send_emails_salon'),
            # sucess 
            path('success/', self.admin_site.admin_view(self.success), name='success'),
        ]
        return custom_urls + urls
#============================================URLS PATHS==============================================
    class Media:
        css = {"all": ("admin/css/forms.css", "css/admin/daterange.css")}
        js = ("admin/js/calendar.js", "js/admin/DateRangeShortcuts.js")

